# app/export/excel_exporter.py

from pathlib import Path
from typing import Any, Optional
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment,
    Border, Side, GradientFill,
)
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter

from app.config import settings
from app.utils.logger import get_logger
from app.utils.helpers import build_export_filename
from app.models.schemas.extracted_data import ExtractedResumeSchema
from app.models.schemas.export_schema import (
    ExportColumnMapping,
    DEFAULT_COLUMN_MAPPING,
)
from app.export.field_mapper import field_mapper

logger = get_logger(__name__)


# ─── Style Constants ───────────────────────────────────────────────────────────
HEADER_BG_COLOR   = "1F3864"   # Dark navy blue
HEADER_FONT_COLOR = "FFFFFF"   # White
ALT_ROW_COLOR     = "EBF0FA"   # Light blue
BORDER_COLOR      = "BDD0E9"   # Soft blue-grey
ACCENT_COLOR      = "2E75B6"   # Medium blue

HIGH_CONF_COLOR   = "C6EFCE"   # Green
MED_CONF_COLOR    = "FFEB9C"   # Yellow
LOW_CONF_COLOR    = "FFC7CE"   # Red


class ExcelExporter:
    """
    Generates formatted Excel (.xlsx) files from extracted resume data.

    Features:
      - Professional styling with header row
      - Alternating row colors
      - Auto-sized columns
      - Confidence score color coding
      - Multiple sheets (Summary, Experience, Education, Skills)
      - Freeze panes and auto-filters
      - Excel Table format for easy filtering
    """

    def __init__(self):
        pass

    # ─── Main Entry ────────────────────────────────────────────────────────────
    def export(
        self,
        schemas:       list[tuple[ExtractedResumeSchema, Optional[dict]]],
        output_path:   Optional[Path] = None,
        mapping:       list[ExportColumnMapping] = DEFAULT_COLUMN_MAPPING,
        include_sheets: bool = True,
    ) -> Path:
        """
        Export multiple resume schemas to a formatted Excel file.

        Args:
            schemas:        List of (schema, metadata) tuples
            output_path:    Where to save the file (auto-generated if None)
            mapping:        Column mapping configuration
            include_sheets: If True, add extra sheets for experience/education

        Returns:
            Path to the generated Excel file
        """
        if output_path is None:
            filename    = build_export_filename("resumes_export", "xlsx")
            output_path = settings.EXPORT_DIR / filename

        logger.info(
            f"Generating Excel export: "
            f"{len(schemas)} resumes → {output_path.name}"
        )

        # Build rows
        rows = field_mapper.map_to_rows(schemas, mapping)

        # Create workbook
        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Resume Summary"
        self._write_summary_sheet(ws_summary, rows, mapping)

        # ── Sheet 2: Experience Detail ────────────────────────────────────────
        if include_sheets:
            ws_exp = wb.create_sheet("Experience Detail")
            self._write_experience_sheet(ws_exp, schemas)

        # ── Sheet 3: Education Detail ──────────────────────────────────────────
        if include_sheets:
            ws_edu = wb.create_sheet("Education Detail")
            self._write_education_sheet(ws_edu, schemas)

        # ── Sheet 4: Skills Matrix ─────────────────────────────────────────────
        if include_sheets:
            ws_skills = wb.create_sheet("Skills Matrix")
            self._write_skills_sheet(ws_skills, schemas)

        # ── Sheet 5: Stats ─────────────────────────────────────────────────────
        if include_sheets:
            ws_stats = wb.create_sheet("Statistics")
            self._write_stats_sheet(ws_stats, schemas)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        logger.info(f"Excel export saved: {output_path}")
        return output_path

    # ─── Summary Sheet ─────────────────────────────────────────────────────────
    def _write_summary_sheet(
        self,
        ws:      Worksheet,
        rows:    list[dict[str, Any]],
        mapping: list[ExportColumnMapping],
    ) -> None:
        """Write the main summary sheet with all candidates."""

        headers = field_mapper.get_headers(mapping)
        widths  = field_mapper.get_column_widths(mapping)

        # ── Title row ─────────────────────────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        title_cell = ws["A1"]
        title_cell.value = (
            f"Resume Extraction Report  |  "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        title_cell.font      = Font(
            name="Calibri", size=13, bold=True,
            color=HEADER_FONT_COLOR
        )
        title_cell.fill      = PatternFill(
            fill_type="solid",
            fgColor=ACCENT_COLOR,
        )
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[1].height = 24

        # ── Header row (row 2) ────────────────────────────────────────────────
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            self._apply_header_style(cell)

        ws.row_dimensions[2].height = 20

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, row_data in enumerate(rows, start=3):
            is_alt = (row_idx % 2 == 0)

            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_data_style(cell, is_alt)

                # Color-code confidence column
                if "confidence" in header.lower() and value:
                    self._apply_confidence_color(cell, value)

            ws.row_dimensions[row_idx].height = 16

        # ── Column widths ──────────────────────────────────────────────────────
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            width      = widths.get(header, 20)
            ws.column_dimensions[col_letter].width = width

        # ── Freeze panes (freeze header rows) ─────────────────────────────────
        ws.freeze_panes = "A3"

        # ── Auto-filter ───────────────────────────────────────────────────────
        last_col = get_column_letter(len(headers))
        last_row = len(rows) + 2
        ws.auto_filter.ref = f"A2:{last_col}{last_row}"

        # ── Excel Table ───────────────────────────────────────────────────────
        if rows:
            table = Table(
                displayName="ResumeData",
                ref=f"A2:{last_col}{last_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

    # ─── Experience Sheet ──────────────────────────────────────────────────────
    def _write_experience_sheet(
        self,
        ws:      Worksheet,
        schemas: list[tuple[ExtractedResumeSchema, Optional[dict]]],
    ) -> None:
        """Write detailed experience sheet."""

        headers = [
            "Full Name", "Email", "Job Title", "Company",
            "Location", "Start Date", "End Date",
            "Duration (Yrs)", "Is Current", "Description",
        ]

        # Write header
        self._write_sheet_header(ws, headers, "Experience Details")

        # Write rows
        row_idx = 3
        for schema, metadata in schemas:
            exp_rows = field_mapper.expand_experience_rows(schema, metadata)
            for exp_row in exp_rows:
                is_alt = (row_idx % 2 == 0)
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=exp_row.get(header, ""),
                    )
                    self._apply_data_style(cell, is_alt)
                row_idx += 1

        self._auto_size_columns(ws, headers)
        ws.freeze_panes = "A3"

    # ─── Education Sheet ───────────────────────────────────────────────────────
    def _write_education_sheet(
        self,
        ws:      Worksheet,
        schemas: list[tuple[ExtractedResumeSchema, Optional[dict]]],
    ) -> None:
        """Write detailed education sheet."""

        headers = [
            "Full Name", "Email", "Degree", "Field of Study",
            "Institution", "Location", "Start Date",
            "Graduation Date", "GPA",
        ]

        self._write_sheet_header(ws, headers, "Education Details")

        row_idx = 3
        for schema, metadata in schemas:
            edu_rows = field_mapper.expand_education_rows(schema, metadata)
            for edu_row in edu_rows:
                is_alt = (row_idx % 2 == 0)
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=edu_row.get(header, ""),
                    )
                    self._apply_data_style(cell, is_alt)
                row_idx += 1

        self._auto_size_columns(ws, headers)
        ws.freeze_panes = "A3"

    # ─── Skills Matrix Sheet ───────────────────────────────────────────────────
    def _write_skills_sheet(
        self,
        ws:      Worksheet,
        schemas: list[tuple[ExtractedResumeSchema, Optional[dict]]],
    ) -> None:
        """
        Write skills matrix sheet.
        Rows = candidates, columns = skills, cells = ✓ or blank.
        """
        # Collect all unique skills
        all_skills: set[str] = set()
        for schema, _ in schemas:
            all_skills.update(schema.skills.all)

        if not all_skills:
            ws["A1"] = "No skills data available"
            return

        skill_list = sorted(all_skills)

        # Headers
        ws.cell(row=1, column=1, value="Full Name")
        ws.cell(row=1, column=2, value="Email")
        for col_idx, skill in enumerate(skill_list, start=3):
            cell = ws.cell(row=1, column=col_idx, value=skill)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        # Apply header style to name/email columns
        self._apply_header_style(ws.cell(row=1, column=1))
        self._apply_header_style(ws.cell(row=1, column=2))
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        # Data rows
        for row_idx, (schema, _) in enumerate(schemas, start=2):
            is_alt = (row_idx % 2 == 0)

            name_cell  = ws.cell(
                row=row_idx, column=1,
                value=schema.contact.full_name or "",
            )
            email_cell = ws.cell(
                row=row_idx, column=2,
                value=schema.contact.email or "",
            )
            self._apply_data_style(name_cell,  is_alt)
            self._apply_data_style(email_cell, is_alt)

            candidate_skills = {s.lower() for s in schema.skills.all}
            for col_idx, skill in enumerate(skill_list, start=3):
                has_skill = skill.lower() in candidate_skills
                cell      = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value="✓" if has_skill else "",
                )
                self._apply_data_style(cell, is_alt)
                if has_skill:
                    cell.font = Font(
                        color="1F7A1F",
                        bold=True,
                    )

        ws.freeze_panes = "C2"

    # ─── Stats Sheet ───────────────────────────────────────────────────────────
    def _write_stats_sheet(
        self,
        ws:      Worksheet,
        schemas: list[tuple[ExtractedResumeSchema, Optional[dict]]],
    ) -> None:
        """Write statistics summary sheet."""

        # Title
        ws.merge_cells("A1:D1")
        title_cell       = ws["A1"]
        title_cell.value = "Extraction Statistics"
        title_cell.font  = Font(
            name="Calibri", size=14, bold=True,
            color=HEADER_FONT_COLOR,
        )
        title_cell.fill      = PatternFill(
            fill_type="solid", fgColor=ACCENT_COLOR
        )
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        # Stats data
        total     = len(schemas)
        with_name = sum(
            1 for s, _ in schemas if s.contact.full_name
        )
        with_email = sum(
            1 for s, _ in schemas if s.contact.email
        )
        with_exp = sum(
            1 for s, _ in schemas if s.experience
        )
        with_edu = sum(
            1 for s, _ in schemas if s.education
        )
        with_skills = sum(
            1 for s, _ in schemas if s.skills.all
        )
        avg_skills = (
            sum(len(s.skills.all) for s, _ in schemas) / total
            if total else 0
        )
        avg_exp = (
            sum(len(s.experience) for s, _ in schemas) / total
            if total else 0
        )
        avg_conf = (
            sum(
                s.confidence_scores.overall or 0
                for s, _ in schemas
            ) / total
            if total else 0
        )

        stats = [
            ("Total Resumes Processed",    total,                     ""),
            ("With Name Extracted",         with_name,                f"{with_name/total*100:.1f}%" if total else ""),
            ("With Email Extracted",        with_email,               f"{with_email/total*100:.1f}%" if total else ""),
            ("With Experience Extracted",   with_exp,                 f"{with_exp/total*100:.1f}%" if total else ""),
            ("With Education Extracted",    with_edu,                 f"{with_edu/total*100:.1f}%" if total else ""),
            ("With Skills Extracted",       with_skills,              f"{with_skills/total*100:.1f}%" if total else ""),
            ("Avg Skills Per Resume",        round(avg_skills, 1),    ""),
            ("Avg Experience Entries",       round(avg_exp, 1),       ""),
            ("Avg Confidence Score",         f"{avg_conf:.1%}",       ""),
            ("Export Generated",             datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
        ]

        # Write stats
        headers = ["Metric", "Value", "Percentage"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            self._apply_header_style(cell)

        for row_idx, (metric, value, pct) in enumerate(stats, start=3):
            ws.cell(row=row_idx, column=1, value=metric).font = Font(
                name="Calibri", size=11,
            )
            ws.cell(row=row_idx, column=2, value=value).font = Font(
                name="Calibri", size=11, bold=True,
            )
            ws.cell(row=row_idx, column=3, value=pct).font = Font(
                name="Calibri", size=11,
            )

            # Alternate row color
            if row_idx % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        fill_type="solid", fgColor=ALT_ROW_COLOR
                    )

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

    # ─── Style Helpers ─────────────────────────────────────────────────────────
    def _apply_header_style(self, cell) -> None:
        """Apply header cell styling."""
        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color=HEADER_FONT_COLOR,
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_BG_COLOR,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
            bottom=Side(style="medium", color=ACCENT_COLOR),
        )

    def _apply_data_style(self, cell, is_alt: bool = False) -> None:
        """Apply data cell styling."""
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=False,
        )

        if is_alt:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ALT_ROW_COLOR,
            )

        cell.border = Border(
            left=Side(style="thin",   color=BORDER_COLOR),
            right=Side(style="thin",  color=BORDER_COLOR),
            bottom=Side(style="thin", color=BORDER_COLOR),
        )

    def _apply_confidence_color(self, cell, value: Any) -> None:
        """Color-code confidence score cells."""
        try:
            score = float(str(value).replace("%", "").strip())

            # Handle percentage format
            if score > 1:
                score = score / 100

            if score >= 0.85:
                color = HIGH_CONF_COLOR
            elif score >= 0.60:
                color = MED_CONF_COLOR
            else:
                color = LOW_CONF_COLOR

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=color,
            )
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=True,
            )
        except (ValueError, TypeError):
            pass

    def _write_sheet_header(
        self,
        ws:      Worksheet,
        headers: list[str],
        title:   str,
    ) -> None:
        """Write a standard title row + header row for a sheet."""

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        title_cell            = ws["A1"]
        title_cell.value      = title
        title_cell.font       = Font(
            name="Calibri", size=13, bold=True,
            color=HEADER_FONT_COLOR,
        )
        title_cell.fill       = PatternFill(
            fill_type="solid", fgColor=ACCENT_COLOR
        )
        title_cell.alignment  = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[1].height = 22

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            self._apply_header_style(cell)
        ws.row_dimensions[2].height = 18

    def _auto_size_columns(
        self,
        ws:      Worksheet,
        headers: list[str],
        min_w:   int = 12,
        max_w:   int = 50,
    ) -> None:
        """Auto-size columns based on content."""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len    = len(str(header))

            for row in ws.iter_rows(
                min_row=3,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

            width = min(max_w, max(min_w, max_len + 2))
            ws.column_dimensions[col_letter].width = width

    # ─── Single Resume Export ──────────────────────────────────────────────────
    def export_single(
        self,
        schema:      ExtractedResumeSchema,
        metadata:    Optional[dict] = None,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Export a single resume to Excel."""
        return self.export(
            schemas      = [(schema, metadata)],
            output_path  = output_path,
            include_sheets=True,
        )


# ─── Singleton ─────────────────────────────────────────────────────────────────
excel_exporter = ExcelExporter()
